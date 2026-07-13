#!/usr/bin/env python3
"""Export Papá's plan (plan_papa.json) + the shared father-son calendar to a
formatted Excel workbook (Spanish, for Papá). Regenerate anytime the plan changes.

  python scripts/plan_to_excel.py            # writes Plan_Papa_UTT59.xlsx
"""
from __future__ import annotations

import json
import sys
from pathlib import Path

RUNNING_DIR = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(RUNNING_DIR))

from openpyxl import Workbook  # noqa: E402
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side  # noqa: E402
from openpyxl.utils import get_column_letter  # noqa: E402

PAPA_OFFSET = 19

PHASE_ES = {"base": "Base", "build": "Construcción", "peak": "Pico",
            "taper": "Afinamiento", "race": "Carrera"}
DAY_ES = {"monday": "Lunes", "tuesday": "Martes", "wednesday": "Miércoles",
          "thursday": "Jueves", "friday": "Viernes", "saturday": "Sábado", "sunday": "Domingo"}
TYPE_ES = {"run": "Carrera fácil", "long_run": "Tirada larga", "descent": "Bajadas",
           "rest": "Descanso", "prehab": "Descanso + prehab", "race": "CARRERA"}

NAVY = "1F3A5F"
GOLD = "E8B923"
GREEN = "3E7D3E"
BLUE = "BBD3E8"
LIGHT = "F2F6FA"
WHITE = "FFFFFF"

thin = Side(style="thin", color="C9D4E0")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)


def header_row(ws, row, headers, fill=NAVY, font_color=WHITE):
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=c, value=h)
        cell.fill = PatternFill("solid", fgColor=fill)
        cell.font = Font(bold=True, color=font_color, size=11)
        cell.alignment = CENTER
        cell.border = BORDER


def row_fill(w):
    if w["phase"] == "race":
        return GREEN, WHITE
    if w["is_recovery"]:
        return BLUE, "000000"
    if w["phase"] == "peak":
        return GOLD, "000000"
    return None, "000000"


def sheet_plan(wb, papa):
    ws = wb.active
    ws.title = "Plan Papá"
    ws.merge_cells("A1:G1")
    t = ws["A1"]
    t.value = "PLAN — Ultra Trail Tarahumara 59 km · 2400 m D+ · 2 oct 2026"
    t.font = Font(bold=True, size=14, color=WHITE)
    t.fill = PatternFill("solid", fgColor=NAVY)
    t.alignment = CENTER
    ws.merge_cells("A2:G2")
    s = ws["A2"]
    s.value = "Meta: terminar fuerte y sano (correr-caminar). 5 días/semana, solo montaña. Tirada larga los sábados (con Emmanuel)."
    s.font = Font(italic=True, size=10)
    s.alignment = CENTER

    headers = ["Sem.", "Fechas", "Fase", "Volumen (km)", "Desnivel (m)", "Tirada larga (km)", "Notas"]
    header_row(ws, 3, headers)
    for i, w in enumerate(papa["weeks"], start=4):
        fill, fc = row_fill(w)
        long_disp = "CARRERA 59 km" if w["phase"] == "race" else f"{w['long_run_km']:.0f}"
        vals = [w["week_number"], f"{w['start_date'][5:]}–{w['end_date'][5:]}",
                PHASE_ES.get(w["phase"], w["phase"]) + (" · Recuperación" if w["is_recovery"] else ""),
                w["distance_km"], w["vert_m"], long_disp, w.get("notes", "")]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(row=i, column=c, value=v)
            cell.border = BORDER
            cell.alignment = LEFT if c == 7 else CENTER
            if fill:
                cell.fill = PatternFill("solid", fgColor=fill)
                cell.font = Font(color=fc, bold=(w["phase"] in ("race", "peak")))
            elif i % 2 == 0:
                cell.fill = PatternFill("solid", fgColor=LIGHT)
    widths = [6, 14, 20, 13, 13, 15, 60]
    for c, wdt in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(c)].width = wdt
    ws.freeze_panes = "A4"


def sheet_detail(wb, papa):
    ws = wb.create_sheet("Detalle semanal")
    header_row(ws, 1, ["Sem.", "Día", "Fecha", "Sesión", "km", "Desnivel (m)", "FC objetivo"])
    r = 2
    for w in papa["weeks"]:
        for wo in w["workouts"]:
            km = wo.get("distance_km")
            vals = [w["week_number"], DAY_ES.get(wo["day"], wo["day"]), wo.get("date", ""),
                    TYPE_ES.get(wo["type"], wo["type"]) + (f" — {wo['description']}" if wo["type"] in ("long_run", "race") else ""),
                    "" if km is None else km, wo.get("vert_m") or "", wo.get("target_hr") or ""]
            for c, v in enumerate(vals, 1):
                cell = ws.cell(row=r, column=c, value=v)
                cell.border = BORDER
                cell.alignment = LEFT if c == 4 else CENTER
                if wo["type"] == "long_run":
                    cell.fill = PatternFill("solid", fgColor="FFF3CC")
                elif wo["type"] == "race":
                    cell.fill = PatternFill("solid", fgColor=GREEN)
                    cell.font = Font(bold=True, color=WHITE)
                elif r % 2 == 0:
                    cell.fill = PatternFill("solid", fgColor=LIGHT)
            r += 1
    for c, wdt in enumerate([6, 11, 12, 42, 7, 12, 14], 1):
        ws.column_dimensions[get_column_letter(c)].width = wdt
    ws.freeze_panes = "A2"


def sheet_shared(wb, papa, emm_plan):
    ws = wb.create_sheet("Juntos (Padre e Hijo)")
    ws.merge_cells("A1:D1")
    t = ws["A1"]
    t.value = "Tirada larga compartida — todos los sábados, misma ruta"
    t.font = Font(bold=True, size=12, color=WHITE)
    t.fill = PatternFill("solid", fgColor=NAVY)
    t.alignment = CENTER
    header_row(ws, 2, ["Fin de semana", "Emmanuel (km)", "Papá (km)", "Juntos"])
    r = 3
    for w in papa["weeks"]:
        n = w["week_number"]
        dates = f"{w['start_date'][5:]}–{w['end_date'][5:]}"
        if n == 12:
            vals = [dates, "CARRERA", "CARRERA", "🏔️ 59 km juntos"]
            fill = GREEN
        else:
            emm = emm_plan.get(n + PAPA_OFFSET)
            el = emm["long_run_km"] if emm else None
            pl = w["long_run_km"]
            if w["is_recovery"] and emm and emm["is_recovery"]:
                nota, fill = "Ambos suave — recuperación", BLUE
            elif el is not None and el >= pl:
                nota, fill = f"🎯 lo alcanzas/superas ({el:.0f} vs {pl:.0f})", GOLD
            else:
                nota, fill = (f"juntos, él sigue +{pl - (el or 0):.0f}", None)
            vals = [dates, f"{el:.0f}" if el is not None else "—", f"{pl:.0f}", nota]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.border = BORDER
            cell.alignment = LEFT if c == 4 else CENTER
            if fill:
                cell.fill = PatternFill("solid", fgColor=fill)
                if fill in (GREEN,):
                    cell.font = Font(bold=True, color=WHITE)
        r += 1
    for c, wdt in enumerate([16, 15, 12, 38], 1):
        ws.column_dimensions[get_column_letter(c)].width = wdt
    ws.freeze_panes = "A3"


def sheet_rules(wb):
    ws = wb.create_sheet("Reglas · Lesiones")
    ws.merge_cells("A1:A1")
    ws["A1"] = "Reglas de oro (prioridad: cuidar de Papá)"
    ws["A1"].font = Font(bold=True, size=12, color=WHITE)
    ws["A1"].fill = PatternFill("solid", fgColor=NAVY)
    rules = [
        "Tope de progresión: no subir el volumen semanal más de ~10% entre semanas de carga.",
        "Semanas de recuperación (2 y 6): bajar ~30%. NO se saltan.",
        "Todo lo fácil, de verdad fácil (~125 ppm). Casi nada de intensidad — así se protege.",
        "Bajadas progresivas: bajadas cortas y controladas al principio; el cuerpo se adapta.",
        "Prehab en casa 2×/semana: sentadilla a una pierna, gemelos, cadera/glúteo (~10 min).",
        "Dolor: si es agudo o de articulación, parar y descansar. Molestia >3 días: bajar carga.",
        "Tirada larga máxima 37 km (por debajo de sus 38 km ya hechos), correr-caminar.",
        "Calor: preparación de calor en las últimas ~3 semanas (la altitud ya está cubierta).",
        "Nutrición: comer/beber cada 40–45 min en las tiradas largas. Proteína ~1.8 g/kg/día.",
    ]
    for i, txt in enumerate(rules, start=2):
        c = ws.cell(row=i, column=1, value=f"• {txt}")
        c.alignment = LEFT
        c.border = BORDER
        if i % 2 == 0:
            c.fill = PatternFill("solid", fgColor=LIGHT)
    ws.column_dimensions["A"].width = 100


def main() -> int:
    from dotenv import load_dotenv
    load_dotenv()
    from tracker import db
    db.init_db()
    papa = json.loads((RUNNING_DIR / "plan_papa.json").read_text())
    emm_plan = {w["week_number"]: w for w in db.get_plan()}

    wb = Workbook()
    sheet_plan(wb, papa)
    sheet_detail(wb, papa)
    sheet_shared(wb, papa, emm_plan)
    sheet_rules(wb)

    out = RUNNING_DIR / "Plan_Papa_UTT59.xlsx"
    wb.save(out)
    print(f"Wrote {out}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
